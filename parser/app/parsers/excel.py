from __future__ import annotations

import io
from datetime import date, datetime

from ..schemas import ParseResult
from .tabular import rows_to_result


class EncryptedFileError(Exception):
    pass


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _is_encrypted_ole2(content: bytes) -> bool:
    # OLE2 (D0CF11E0) contendo o stream "EncryptionInfo" => Office criptografado
    return content[:4] == b"\xd0\xcf\x11\xe0" and b"E\x00n\x00c\x00r\x00y\x00p\x00t\x00i\x00o\x00n" in content[:4096]


def parse_xlsx(content: bytes) -> ParseResult:
    if content[:4] == b"\xd0\xcf\x11\xe0":
        # extensão .xlsx mas conteúdo OLE2 — na verdade é .xls
        return parse_xls(content)
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("openpyxl não instalado") from e

    # read_only=False: o modo streaming da openpyxl trunca alguns extratos de
    # banco (confia numa dimensão declarada errada). Planilha de extrato é pequena.
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True, keep_links=False)
    ws = wb.active
    rows = [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows_to_result(rows, "xlsx")


def parse_xls(content: bytes) -> ParseResult:
    if _is_encrypted_ole2(content):
        raise EncryptedFileError(
            "planilha protegida por senha — mande o extrato em OFX ou CSV, "
            "ou remova a senha da planilha antes de subir"
        )
    try:
        import xlrd
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("xlrd não instalado") from e

    try:
        book = xlrd.open_workbook(file_contents=content)
    except xlrd.XLRDError as e:
        # "Can't find workbook in OLE2" costuma ser arquivo protegido/corrompido
        raise EncryptedFileError(
            f"não consegui abrir a planilha ({e}). Se ela tem senha, mande o OFX/CSV."
        ) from e

    sheet = book.sheet_by_index(0)
    rows: list[list[str]] = []
    for r in range(sheet.nrows):
        out = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                y, mo, d, *_ = xlrd.xldate_as_tuple(cell.value, book.datemode)
                out.append(f"{d:02d}/{mo:02d}/{y:04d}")
            else:
                out.append(_cell_str(cell.value))
        rows.append(out)
    return rows_to_result(rows, "xls")
